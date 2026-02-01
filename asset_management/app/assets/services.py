import csv
from datetime import datetime
from io import StringIO, BytesIO
from typing import Annotated, List

from fastapi import Depends, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from asset_management.app.assets.repositories import AssetRepository
from asset_management.app.assets.schemas import AssetCreateRequest, AssetResponse, AssetUpdateRequest
from asset_management.app.assets.models import Asset


class AssetService:
    def __init__(
        self,
        asset_repository: Annotated[AssetRepository, Depends()],
    ) -> None:
        self.asset_repository = asset_repository

    def create_asset_for_admin(
        self, admin_club_id: int, asset_request: AssetCreateRequest
    ) -> AssetResponse:
        
        new_asset = Asset(
            name=asset_request.name,
            description=asset_request.description,
            club_id=admin_club_id,
            category_id=asset_request.category_id,
            total_quantity=asset_request.quantity,
            available_quantity=asset_request.quantity,
            location=asset_request.location,
            max_rental_days=asset_request.max_rental_days,
        )

        self.asset_repository.create_asset(new_asset)

        return AssetResponse(
            id=new_asset.id,
            name=new_asset.name,
            status=0,
            description=new_asset.description,
            club_id=new_asset.club_id,
            category_id=new_asset.category_id,
            category_name=None,  # To be filled if needed
            total_quantity=new_asset.total_quantity,
            available_quantity=new_asset.available_quantity,
            location=new_asset.location,
            created_at=new_asset.created_at,
            max_rental_days=new_asset.max_rental_days,
        )

    def update_asset_for_admin(
        self, admin_club_id: int, asset_id: int, asset_request: AssetUpdateRequest
    ) -> AssetResponse:
        
        asset = self.asset_repository.get_asset_by_id(asset_id)
        if asset is None:
            raise Exception("ItemNotFoundException")  # Replace with proper exception

        updated_asset = self.asset_repository.modify_asset(
            asset,
            name=asset_request.name,
            description=asset_request.description,
            club_id=admin_club_id,
            category_id=asset_request.category_id,
            total_quantity=asset_request.quantity,
            available_quantity=asset_request.quantity,
            location=asset_request.location,
            max_rental_days=asset_request.max_rental_days,
        )

        return AssetResponse(
            id=updated_asset.id,
            name=updated_asset.name,
            description=updated_asset.description,
            status=self.asset_repository.get_asset_status(updated_asset.id),
            category_id=updated_asset.category_id,
            category_name=None,  # To be filled if needed
            total_quantity=updated_asset.total_quantity,
            available_quantity=updated_asset.available_quantity,
            location=updated_asset.location,
            created_at=updated_asset.created_at,
            max_rental_days=updated_asset.max_rental_days,
        )

    def delete_asset_for_admin(self, asset_id: int) -> None:

        asset = self.asset_repository.get_asset_by_id(asset_id)
        if asset is None:
            raise Exception("ItemNotFoundException")  # Replace with proper exception
        
        self.asset_repository.delete_asset(asset)

    def list_assets_for_club(self, club_id: int) -> List[AssetResponse]:
        assets = self.asset_repository.get_all_assets_in_club(club_id)
        return [
            AssetResponse(
                id=asset.id,
                name=asset.name,
                status=self.asset_repository.get_asset_status(asset.id),
                description=asset.description,
                club_id=asset.club_id,
                category_id=asset.category_id,
                category_name=None,  # To be filled if needed
                total_quantity=asset.total_quantity,
                available_quantity=asset.available_quantity,
                location=asset.location,
                created_at=asset.created_at,
                max_rental_days=asset.max_rental_days,
            )
            for asset in assets
        ]
    
    def generate_import_template(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Asset Template"
        
        # 헤더 작성
        headers = ["name", "description", "total_quantity", "available_quantity", "location", "created_at"]
        ws.append(headers)
        
        # 예시 행 작성
        ws.append(["예시 행입니다. 이름은 50자", "설명은 500자", "최대 수량", "대여 가능 수량", "동아리 내 위치(100자 이하)", "2024-01-01 00:00:00"])
        
        # 열 너비 자동 조정
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    async def import_assets_from_excel(self, club_id: int, file: UploadFile) -> dict:
        contents = await file.read()
        
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        failed = []
        imported_count = 0
        
        # 헤더 행 가져오기 (첫 번째 행)
        headers = [cell.value for cell in ws[1]]
        
        # 데이터 행 처리 (2번째 행부터)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 빈 행 건너뛰기
                continue
            try:
                row_dict = dict(zip(headers, row))
                self.asset_repository.create_asset(Asset(
                    name=str(row_dict["name"]),
                    description=str(row_dict["description"]) if row_dict["description"] else "",
                    total_quantity=int(row_dict["total_quantity"]),
                    available_quantity=int(row_dict["available_quantity"]),
                    location=str(row_dict["location"]) if row_dict["location"] else "",
                    created_at=datetime.strptime(str(row_dict["created_at"]), "%Y-%m-%d %H:%M:%S"),
                    club_id=club_id
                ))
                imported_count += 1
            except Exception as e:
                failed.append({"row": row_dict if 'row_dict' in locals() else row, "error": str(e)})
        return {"imported": imported_count, "failed": failed}
        
        
    
    def export_assets_to_excel(self, club_id: int) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Assets"
        
        # 헤더 작성
        headers = ["name", "description", "total_quantity", "available_quantity", "location", "created_at"]
        ws.append(headers)
        
        # 자산 데이터 작성
        assets = self.asset_repository.get_all_assets_in_club(club_id)
        for asset in assets:
            ws.append([
                asset.name,
                asset.description,
                asset.total_quantity,
                asset.available_quantity,
                asset.location,
                asset.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ])
        
        # 열 너비 자동 조정
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()